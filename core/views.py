from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .forms import LoginForm, ProductoForm
from .models import Producto, Categoria, Boleta, Conversacion, Mensaje, Tamano, Unidad_medida, VariantePrecio, Ticket
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.views.decorators.csrf import csrf_exempt
import uuid
from django.core.files.base import ContentFile
from openpyxl import load_workbook
import barcode
from barcode import generate
from barcode.writer import ImageWriter
from io import BytesIO
from django.http import HttpResponse
from django.core.files.base import ContentFile


# Create your views here.


def home(request):
    return render(request, "core/home.html")


def crear(request):
    if request.method == 'POST':
        value = request.POST.get('barcode_value', '')

        # Genera el código de barras en formato CODE128
        barcode_class = barcode.get_barcode_class('code128')
        code = barcode_class(value, writer=ImageWriter())
        buffer = BytesIO()
        code.write(buffer)
        code_img = ContentFile(buffer.getvalue())

        context = {'code_img': code_img}
        return render(request, 'core/crear.html', context)

    return render(request, 'core/crear.html')


def login(request):
    return render(request, "core/login.html")


def login_view(request):
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            password = form.cleaned_data["password"]
            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                # Redirige a la página deseada después del inicio de sesión
                return redirect("pagina_despues_login")
    else:
        form = LoginForm()
    return render(request, "login.html", {"form": form})


def logout_view(request):
    logout(request)
    # Redirige a la página deseada después del cierre de sesión
    return redirect("pagina_despues_logout")

@login_required
def index(request):
    return render(request, "core/index.html")

from django.core.paginator import Paginator

@login_required
def listado_productos(request):
    productos = Producto.objects.all()
    categorias = Categoria.objects.all()

    # Obtener los valores de los filtros enviados por el formulario
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    categorias_seleccionadas = request.GET.getlist('categoria')
    imagen_predeterminada_url = '/media/productos/default.jpg'

    # Filtrar por precio mínimo y máximo
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)

    # Filtrar por categorías seleccionadas
    if categorias_seleccionadas:
        productos = productos.filter(categoria__in=categorias_seleccionadas)

    # Agregar paginador
    paginator = Paginator(productos, 16)  # Mostrar 10 productos por página
    page_number = request.GET.get('page')
    productos = paginator.get_page(page_number)

    for producto in productos:
        if not producto.imagen:
            producto.imagen = imagen_predeterminada_url

    context = {
        'productos': productos,
        'categorias': categorias
    }
    return render(request, 'core/listado_productos.html', context)


@login_required
def nuevo_producto(request):
    data = {"form": ProductoForm()}
    if request.method == 'POST':
        formulario = ProductoForm(request.POST)
        if formulario.is_valid():
            formulario.save()
            data['mensaje'] = "Guardado Correctamente"
    return render(request, "core/nuevo_producto.html", data)


@login_required
def modificar_producto(request, id):
    producto = Producto.objects.get(id=id)
    data = {
        'form' :ProductoForm(instance=producto)
    }

    if request.method == 'POST':
        formulario = ProductoForm(data=request.POST, instance=producto)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "El producto fue modificado Correctamente")
            return redirect("modificar")
    return render(request, 'core/modificar_producto.html', data)
@login_required
def eliminar_producto(request, id):
    try:
        producto = Producto.objects.get(id=id)
        producto.delete()
        messages.success(request, "El producto fue eliminado Correctamente")
        return redirect("modificar")
    except Producto.DoesNotExist:
        messages.error(request, "El producto no existe")
        return redirect("modificar")

@login_required
def agregar(request):
    return render(request, 'core/agregar.html')

@login_required
def buscar_producto(request):
    search_input = request.GET.get('search_input', '').lower()
    # Filtrar los productos que coincidan con el criterio de búsqueda
    filtered_products = Producto.objects.filter(nombre__icontains=search_input)

    context = {
        'products': filtered_products,
        'search_input': search_input,
    }
    return render(request, 'core/buscar_producto.html', context)
@login_required
def ventas(request):
    return render(request, 'core/ventas.html')
@login_required
def buscar_producto_ajax(request):
    search_input = request.GET.get('search_input', '').lower()
    # Filtrar los productos que coincidan con el criterio de búsqueda
    filtered_products = Producto.objects.filter(nombre__icontains=search_input)

    # Preparar los resultados para enviar como respuesta JSON
    results = [
        {
            'nombre': product.nombre,
            'descripcion': product.descripcion,
            'precio': str(product.precio),
            'categoria': product.categoria.nombre,
        }
        for product in filtered_products
    ]
    return JsonResponse(results, safe=False)
def generar_boleta(request):
    usuario = request.user
    # Obtener los IDs de los productos seleccionados en el carrito desde la solicitud del usuario
    shopping_cart_ids = request.POST.getlist('product_ids')  # Obtener los IDs desde el formulario o solicitud

    # Obtener los productos correspondientes a los IDs del carrito
    shoppingCart = [get_object_or_404(Producto, id=product_id) for product_id in shopping_cart_ids]

    # Calcular el total a pagar
    total_amount = sum(product.precio * (1 - product.descuento / 100) for product in shoppingCart)

    # Generar un identificador único para el nombre del archivo PDF
    unique_id = uuid.uuid4().hex
    file_name = f"boleta_{unique_id}.pdf"

    # Generar el PDF de la boleta usando reportlab.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    p = canvas.Canvas(response, pagesize=letter)
    p.drawString(100, 750, 'BOLETA DE VENTA')

    y_position = 700
    for product in shoppingCart:
        p.drawString(100, y_position, f"{product.nombre} - Precio: ${product.precio} - Descuento: {product.descuento}%")
        y_position -= 20

    p.drawString(100, y_position, f'Total a Pagar: ${total_amount:.2f}')

    p.save()

    # Crear una instancia de Boleta y guardarla en la base de datos
    boleta = Boleta(total_a_pagar=total_amount, nro_boleta=unique_id, usuario=usuario)
    boleta.archivo_pdf.save(file_name, ContentFile(response.content))  
    return response



def agregar_al_carrito(request, producto_id):
    producto = Producto.objects.get(pk=producto_id)

    if 'carrito' not in request.session:
        request.session['carrito'] = []

    request.session['carrito'].append(producto_id)
    request.session.modified = True

    return redirect('carrito_ventas') 


@login_required
def modificar(request):
    productos = Producto.objects.all()
    categorias = Categoria.objects.all()

    # Obtener los valores de los filtros enviados por el formulario
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    categorias_seleccionadas = request.GET.getlist('categoria')

    # Filtrar por precio mínimo y máximo
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)

    # Filtrar por categorías seleccionadas
    if categorias_seleccionadas:
        productos = productos.filter(categoria__in=categorias_seleccionadas)

    # Agregar paginador
    paginator = Paginator(productos, 16)  # Mostrar 16 productos por página
    page_number = request.GET.get('page')
    productos_paginados = paginator.get_page(page_number)

    return render(request, 'core/modificar.html', {'productos': productos_paginados, 'categorias': categorias})

# firebase login
# firebase init
# firebase deploy
@login_required
def carrito_ventas(request):
    carrito_productos = []
    carrito_ids = request.session.get('carrito', [])

    for producto_id in carrito_ids:
        try:
            producto = Producto.objects.get(pk=producto_id)
            carrito_productos.append(producto)
        except Producto.DoesNotExist:
            # Manejar el caso en el que el producto no existe
            pass  # Puedes mostrar un mensaje de error o simplemente omitir el producto

    return render(request, 'core/carrito_ventas.html', {'carrito_productos': carrito_productos})



def realizar_pago(request):
    # Aquí iría la lógica para procesar el pago, por ejemplo, interactuar con una pasarela de pago.
    # Dependiendo de tu implementación, puede ser necesario usar formularios, recibir datos de la solicitud,
    # procesar la información del carrito de compras, calcular el total, etc.

    # Ejemplo básico: supongamos que el pago es exitoso y queremos mostrar un mensaje de éxito.
    mensaje = "¡Pago realizado con éxito! Gracias por tu compra."

    return render(request, 'core/pago_exitoso.html', {'mensaje': mensaje})


def chat_room(request):
    return render(request, 'core/chat.html')

# Vista para manejar el envío de mensajes
def send_message(request):
    if request.method == 'POST':
        message = request.POST.get('message')  # Obtener el mensaje del formulario

        # Aquí debes implementar la lógica para enviar el mensaje al personal de la empresa o al usuario que lo necesita.
        # Puedes utilizar Django Channels o WebSockets para lograr la comunicación en tiempo real.

        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'})

# Vista para obtener mensajes anteriores
def get_messages(request):
    # Aquí debes implementar la lógica para obtener los mensajes anteriores desde la base de datos o cualquier otra fuente.
    messages = [
        {'sender': 'usuario', 'content': 'Hola, necesito ayuda'},
        {'sender': 'personal', 'content': '¡Hola! ¿En qué puedo ayudarte?'}
    ]

    return JsonResponse({'messages': messages})

@login_required

def worker_chat(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('chat-room')
        else:
            # Mostrar un mensaje de error o realizar otras acciones en caso de autenticación fallida.

            return render(request, 'worker_chat.html')

# views.py
def evaluation(request):
    if request.method == 'POST':
        rating = request.POST.get('rating')
        feedback = request.POST.get('feedback')

        # Aquí debes implementar la lógica para almacenar la evaluación en la base de datos o realizar otras acciones necesarias.

        return redirect('chat-room')  # Redirigir a la página de chat después de enviar la evaluación.

    return render(request, 'evaluation.html')
@login_required
def nueva_conversacion(request):
    # Aquí puedes implementar la lógica para crear una nueva conversación en la base de datos.
    # Esto podría implicar la creación de un nuevo registro en tu modelo Conversacion.

    # Ejemplo: Crear una nueva conversación en la base de datos
    nueva_conversacion = Conversacion.objects.create()

    # Lógica adicional aquí si es necesario

    # Redirigir a la página de chat después de crear la conversación.
    return redirect('chat-room')  # Asegúrate de que 'chat-room' sea la URL correcta de tu página de chat.


@login_required
def send_message(request, conversation_id):
    if request.method == 'POST':
        message_content = request.POST.get('message')

        # Obtener la conversación asociada al ID
        conversation = get_object_or_404(Conversacion, id=conversation_id)

        # Crear un nuevo mensaje
        nuevo_mensaje = Mensaje.objects.create(
            conversacion=conversation,
            contenido=message_content,
            remitente=request.user
        )

        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'})


@login_required
def get_messages(request, conversation_id):
    if request.method == 'GET':
        conversation_id = int(request.GET.get('conversation_id'))

        # Obtener la conversación y sus mensajes asociados
        conversation = get_object_or_404(Conversacion, id=conversation_id)
        mensajes = Mensaje.objects.filter(conversacion=conversation)

        # Construir una lista de mensajes
        messages_list = []
        for mensaje in mensajes:
            messages_list.append({
                'sender': mensaje.remitente.username,
                'content': mensaje.contenido,
                'timestamp': mensaje.fecha_envio.strftime('%Y-%m-%d %H:%M:%S')
            })

        return JsonResponse({'messages': messages_list})
    else:
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'})

def cargar_productos_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active
        default_categoria = Categoria.objects.get(nombre='Detergente')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_barras = row[0]
            nombre = row[1]
            precio = row[2]
            stock = row[3]
            descuento = row[4] or 0

            # Crear un nuevo objeto Producto y guardarlo en la base de datos
            producto = Producto(
                codigo_barras=codigo_barras,
                nombre=nombre,
                precio=precio,
                stock=stock,
                descuento=descuento,
                categoria=default_categoria
            )
            producto.save()

        return redirect('listado_productos') 

    return render(request, 'core/cargar_productos.html')



def actualizar_informacion_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_barras = row[0]
            nuevo_nombre = row[1]
            nuevo_precio = row[2]
            nuevo_stock = row[3]
            nuevo_descuento = row[4] or 0
            try:
                # Obtener el producto existente por código de barras
                producto = Producto.objects.get(codigo_barras=codigo_barras)
                # Actualizar el nombre y el precio del producto
                producto.nombre = nuevo_nombre
                producto.precio = nuevo_precio
                producto.stock = nuevo_stock
                producto.descuento =nuevo_descuento
                producto.save()
            except Producto.DoesNotExist:
                pass  # Ignorar productos que no existen en la base de datos

        return redirect('listado_productos')  # Redirigir a la lista de productos

    return render(request, 'actualizar_informacion_desde_excel.html')


def cargar_categorias_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_categoria = row[0]

            # Crear un nuevo objeto Categoria y guardarlo en la base de datos
            categoria = Categoria(nombre=nombre_categoria)
            categoria.save()

        return redirect('listado_productos') 

    return render(request, 'core/cargar_categorias.html')


def cargar_unidades_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_unidad = row[0]

            # Crear un nuevo objeto Unidad_medida y guardarlo en la base de datos
            unidad = Unidad_medida(nombre=nombre_unidad)
            unidad.save()

        return redirect('listado_productos') 

    return render(request, 'core/cargar_unidades.html')


def cargar_tamanos_desde_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Cargar el archivo Excel
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre_tamano = row[0]

            # Crear un nuevo objeto Tamano y guardarlo en la base de datos
            tamano = Tamano(nombre=nombre_tamano)
            tamano.save()

        return redirect('listado_productos') 

    return render(request, 'core/cargar_tamanos.html')


def escanear_codigo_barras(request):
    codigo_barras = request.GET.get('codigo_barras', None)
    producto = None
    variantes = []

    if codigo_barras:
        try:
            producto = Producto.objects.get(codigo_barras=codigo_barras)
            variantes = VariantePrecio.objects.filter(producto=producto)
        except Producto.DoesNotExist:
            pass

    return render(request, 'core/escanear_codigo_barras.html', {'producto': producto, 'variantes': variantes})



def soporte(request):
    if request.method == 'POST':
        name = request.POST['name']
        email = request.POST['email']
        phone_number = request.POST['phone_number']
        message = request.POST['message']
        # Crea el ticket en la base de datos
        ticket = Ticket.objects.create(name=name, email=email, phone_number=phone_number, message=message)
        # Redirige a donde desees después de crear el ticket
        return redirect('core/home')
    return render(request, 'core/soporte.html')


def create_ticket(request):
    tickets = Ticket.objects.all()
    context = {'tickets': tickets}
    return render(request, "core/create_ticket.html", context)

def responder_ticket(request, ticket_id):
    ticket = Ticket.objects.get(id=ticket_id)
    if request.method == 'POST':
        respuesta = request.POST['respuesta']
        estado = request.POST['estado']
        
        # Actualiza el campo de respuesta y el estado en el ticket
        ticket.respuesta = respuesta
        ticket.status = estado
        ticket.save()
        
        return redirect('create_ticket')  # Ajusta la redirección según tu necesidad
    
    context = {'ticket': ticket}
    return render(request, 'core/responder_ticket.html', context)